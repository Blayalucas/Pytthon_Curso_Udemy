# OPenpyxl com Cell

"""
Ela chama a planilha igual no excel com cabecalho e linhas
    
"""

# biblioteca para atualizar
from pathlib import Path

# importar a biblioteca openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell


# Caminho para a raiz do projeto
ROOT_FOLDER = Path(__file__).parent
# O arquivo do excel
WORKBOOK_PATH = ROOT_FOLDER / 'Financeiro.xlsx'

# chamando o arquivo
workbook: Workbook = load_workbook(WORKBOOK_PATH)


# Nome da planilha 

sheet_name = "Financeiro 2024"

# Selecionando  a planilha

worksheet: Worksheet = workbook[sheet_name]


# loop for CEll


rom:tuple[Cell]
for row in worksheet.iter_rows ():
    for cell in row:
        print(cell.value, end='\t')
    print()
    


# Salvando as informações

workbook.save(WORKBOOK_PATH)
    

