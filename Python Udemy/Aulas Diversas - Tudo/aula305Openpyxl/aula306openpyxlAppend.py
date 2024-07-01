# Openpyxl com for Append

"""
Esse for para criar um loop com adicionamento de linhas


"""
# biblioteca para atualizar
from pathlib import Path

# importar a biblioteca openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


# Caminho para a raiz do projeto
ROOT_FOLDER = Path(__file__).parent
# O arquivo do excel
WORKBOOK_PATH = ROOT_FOLDER / 'Financeiro.xlsx'

# chamando o arquivo
workbook = Workbook()
worksheet: Worksheet = workbook.active


# Criando cabeçalho " linha" " coluna" e o "nome da coluna"

worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')


# Instancia de colunas igual excel
students = [

    ['joao', 14, 5.5],
    ["Maria", 16, 7.5],
    ["Pedro", 17, 4.5],

]

# loop for com append

for student in students:
    worksheet.append(student)


# Salvando as informações

workbook.save(WORKBOOK_PATH)
