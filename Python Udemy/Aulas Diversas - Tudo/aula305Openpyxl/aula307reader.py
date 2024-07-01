# OPenpyxl com reader 

# biblioteca para atualizar
from pathlib import Path

# importar a biblioteca openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# Caminho para a raiz do projeto
ROOT_FOLDER = Path(__file__).parent
# O arquivo do excel
WORKBOOK_PATH = ROOT_FOLDER / 'Financeiro.xlsx'

# chamando o arquivo
workbook: Workbook = load_workbook(WORKBOOK_PATH)


# Nome da planilha 

sheet_name = "Financeiro 2024"

# Selecionando  a planilha

worksheet: Worksheet = workbook[sheet_name]


# loop for 

for row in worksheet.iter_rows ():
    for column in rows:
        print(column)
    


# Salvando as informações

workbook.save(WORKBOOK_PATH)
