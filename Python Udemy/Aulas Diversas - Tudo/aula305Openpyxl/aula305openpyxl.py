# Openpyxl 

"""
É uma biblioteca exclusiva para trabalhar com arquivos EXCEL xlsx

Com ela é possivel ler e escrever dados em células especioficas, formatar celulas 
inserir graficos, criar formulas, adicionar imagens e outros elementos gráficas às suas 
planilhas. Ela é util para automatizar tarefas envolvendo planilhas manipulação 
de grandes quantidades de informações 

Observação : Necessario instalar pip install openpyxl

"""
# biblioteca para atualizar 
from pathlib import Path

# importar a biblioteca openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


# Caminho para a raiz do projeto
ROOT_FOLDER = Path(__file__).parent
# O arquivo do excel  
WORKBOOK_PATH = ROOT_FOLDER / 'Financeiro.xlsx'

# chamando o arquivo 
workbook = Workbook()
worksheet: Worksheet = workbook.active


# Criando cabeçalho " linha" " coluna" e o "nome da coluna"

worksheet.cell(1,1,'Nome')
worksheet.cell(1,2,'Idade')
worksheet.cell(1,3,'Nota')


# Instancia de colunas igual excel
students = [
    
    ['joao', 14, 5.5],
    ["Maria", 16, 7.5],
    ["Pedro", 17, 4.5],
    
]




# Salvando as informações 

workbook.save(WORKBOOK_PATH)